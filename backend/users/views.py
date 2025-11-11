from rest_framework import generics, status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
from .serializers import (
    UserSerializer, 
    UserRegistrationSerializer, 
    UserUpdateSerializer,
    ChangePasswordSerializer,
    UserPermissionSerializer,
    BulkUserPermissionSerializer,
    CustomTokenObtainPairSerializer,
    UserLocationMappingSerializer,
    BulkUserLocationMappingSerializer,
)
from .models import UserPermission, MenuItemType, GroupPermission

# Import POS Function models and serializers conditionally
try:
    from .models import POSFunction, RolePOSFunctionMapping
    # Import serializers after models are imported
    # Serializers are conditionally defined, so we need to check if they exist
    from . import serializers as users_serializers
    POSFunctionSerializer = getattr(users_serializers, 'POSFunctionSerializer', None)
    RolePOSFunctionMappingSerializer = getattr(users_serializers, 'RolePOSFunctionMappingSerializer', None)
except (ImportError, AttributeError) as e:
    POSFunction = None
    RolePOSFunctionMapping = None
    POSFunctionSerializer = None
    RolePOSFunctionMappingSerializer = None
from django.contrib.auth.models import Group
from .role_permissions import apply_role_template_to_user, get_role_permissions

User = get_user_model()


class CustomTokenObtainPairView(TokenObtainPairView):
    """
    Custom token view that uses CustomTokenObtainPairSerializer
    to ensure proper authentication with custom User model.
    """
    serializer_class = CustomTokenObtainPairSerializer
    permission_classes = [AllowAny]  # Allow unauthenticated access for login


class RegisterView(generics.CreateAPIView):
    """
    Register a new user.
    
    Anyone can register (no authentication required).
    """
    serializer_class = UserRegistrationSerializer
    permission_classes = [permissions.AllowAny]
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        return Response(
            {
                'message': 'User registered successfully',
                'user': UserSerializer(user).data
            },
            status=status.HTTP_201_CREATED
        )


class CurrentUserView(APIView):
    """
    Get the current authenticated user's information.
    """
    serializer_class = UserSerializer
    
    def get(self, request):
        serializer = UserSerializer(request.user)
        return Response(serializer.data)


class UpdateProfileView(generics.UpdateAPIView):
    """
    Update the current user's profile information.
    """
    serializer_class = UserUpdateSerializer
    
    def get_object(self):
        return self.request.user


class UserDetailUpdateView(generics.RetrieveUpdateAPIView):
    """
    Retrieve or update a specific user by ID.
    Admin can update any user, users can only update themselves via profile endpoint.
    """
    queryset = User.objects.all()
    serializer_class = UserSerializer
    
    def get_serializer_class(self):
        """Use update serializer for PATCH requests"""
        if self.request.method in ['PATCH', 'PUT']:
            return UserUpdateSerializer
        return UserSerializer
    
    def get_queryset(self):
        user = self.request.user
        
        # Only admin and superuser can update other users
        if user.is_superuser or user.role == 'admin':
            return User.objects.all()
        
        # Others can only see/update themselves
        return User.objects.filter(id=user.id)


class ChangePasswordView(APIView):
    """
    Change the current user's password.
    """
    serializer_class = ChangePasswordSerializer
    
    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        user = request.user
        
        # Check old password
        if not user.check_password(serializer.validated_data['old_password']):
            return Response(
                {'old_password': 'Wrong password.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Set new password
        user.set_password(serializer.validated_data['new_password'])
        user.save()
        
        return Response(
            {'message': 'Password changed successfully'},
            status=status.HTTP_200_OK
        )


class UserListView(generics.ListAPIView):
    """
    List all users.
    - Superuser: No restrictions, can see all users
    - Admin role: Can see all users (for permission management)
    - Other authenticated users: Can see all users (needed for permission matrix UI display)
    
    Note: Permission changes are restricted to admin role only (enforced in Permission APIs).
    """
    queryset = User.objects.all()
    serializer_class = UserSerializer
    
    def get_queryset(self):
        user = self.request.user
        
        # Superuser: No restrictions, can see all users
        if user.is_superuser:
            return User.objects.all()
        
        # Admin role: Can see all users for permission management
        if user.role == 'admin':
            return User.objects.all()
        
        # For permission matrix UI, allow all authenticated users to see user list
        # But restrict actual permission changes to admin role only (enforced in Permission APIs)
        return User.objects.all()


# Permission Views
class UserPermissionViewSet(ModelViewSet):
    """
    CRUD operations for user permissions.
    Only admin role can manage permissions.
    """
    serializer_class = UserPermissionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        user_id = self.kwargs.get('user_pk')
        
        # Only admin role can manage permissions
        if user.role != 'admin' and not user.is_superuser:
            return UserPermission.objects.none()
        
        if user_id:
            return UserPermission.objects.filter(user_id=user_id)
        return UserPermission.objects.all()
    
    def create(self, request, *args, **kwargs):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can manage permissions'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can manage permissions'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can manage permissions'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)


class BulkUserPermissionView(APIView):
    """
    Bulk update user permissions.
    Only admin role can perform bulk updates.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can manage permissions'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = BulkUserPermissionSerializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        
        created_count = 0
        updated_count = 0
        
        for item in serializer.validated_data:
            user_id = item['user_id']
            permissions_data = item['permissions']
            
            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                continue
            
            for menu_item_id, perms in permissions_data.items():
                try:
                    menu_item = MenuItemType.objects.get(menu_item_id=menu_item_id)
                except MenuItemType.DoesNotExist:
                    continue
                
                permission, created = UserPermission.objects.update_or_create(
                    user=user,
                    menu_item=menu_item,
                    defaults={
                    'can_access': perms.get('can_access', False),
                    'can_view': perms.get('can_view', False),
                    'can_create': perms.get('can_create', False),
                    'can_edit': perms.get('can_edit', False),
                    'can_delete': perms.get('can_delete', False),
                        'override': perms.get('override', False),
                        'created_by': request.user,
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
        
        return Response({
            'message': 'Permissions updated successfully',
            'created': created_count,
            'updated': updated_count,
        }, status=status.HTTP_200_OK)


class GetUserPermissionsView(APIView):
    """
    Get all permissions for a specific user.
    Returns permissions in format suitable for frontend matrix.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, user_pk=None, user_id=None):
        # Handle both user_pk (from URL) and user_id (for compatibility)
        target_user_id = user_pk or user_id
        
        # If user_id not provided, return current user's permissions
        if not target_user_id:
            target_user_id = request.user.id
        
        # Check if requesting user has permission
        requesting_user = request.user
        try:
            target_user = User.objects.get(id=target_user_id)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Users can see their own permissions, admin can see all
        if requesting_user.id != target_user.id and requesting_user.role != 'admin' and not requesting_user.is_superuser:
            return Response(
                {'error': 'Permission denied'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get all permissions for user
        try:
            user_permissions = UserPermission.objects.filter(user_id=target_user_id).select_related('menu_item')
            
            # Format for frontend
            permissions_dict = {}
            for perm in user_permissions:
                # Skip if menu_item is None (orphaned permission)
                if not perm.menu_item:
                    continue
                
                permissions_dict[perm.menu_item.menu_item_id] = {
                    'can_view': perm.can_view,
                    'can_create': perm.can_create,
                    'can_edit': perm.can_edit,
                    'can_delete': perm.can_delete,
                    'override': perm.override,
                }
            
            return Response(permissions_dict, status=status.HTTP_200_OK)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error fetching user permissions for user {user_id}: {str(e)}")
            return Response(
                {'error': f'Error fetching permissions: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ApplyRoleTemplateView(APIView):
    """
    Apply role template permissions to a user.
    Only admin role can apply templates.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can apply role templates'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        user_id = request.data.get('user_id')
        role_template = request.data.get('role_template')  # Optional: specific role template to apply
        
        if not user_id:
            return Response(
                {'error': 'user_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Apply role template (use provided template or user's current role)
        template_permissions = apply_role_template_to_user(user, role_template=role_template)
        
        created_count = 0
        for menu_item_id, perms in template_permissions.items():
            try:
                menu_item = MenuItemType.objects.get(menu_item_id=menu_item_id)
            except MenuItemType.DoesNotExist:
                continue
            
            permission, created = UserPermission.objects.update_or_create(
                user=user,
                menu_item=menu_item,
                defaults={
                    'can_access': perms.get('can_access', False),
                    'can_view': perms.get('can_view', False),
                    'can_create': perms.get('can_create', False),
                    'can_edit': perms.get('can_edit', False),
                    'can_delete': perms.get('can_delete', False),
                    'override': False,  # Template permissions are not overrides
                    'created_by': request.user,
                }
            )
            
            if created:
                created_count += 1
        
        return Response({
            'message': f'Role template applied successfully',
            'permissions_created': created_count,
        }, status=status.HTTP_200_OK)


class GetRoleTemplateView(APIView):
    """
    Get role template details for viewing.
    Returns template information including all menu items and their permissions.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, role_template):
        # Check admin permission (optional - could allow all authenticated users to view templates)
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can view role templates'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get template for the role
        template = get_role_permissions(role_template)
        
        if not template:
            return Response(
                {'error': f'Template not found for role: {role_template}'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # List of hidden menu items (Archive and POS Phase 2 items)
        # These items are hidden from the UI but may still exist in role templates
        # They will be filtered out from the template view dialog
        HIDDEN_MENU_ITEMS = {
            # Archive items (Users, Security)
            'users',
            'security',
            
            # POS Phase 2 items (all items with parentCategory='pos_phase_2' - hidden from UI)
            'pos_phase_2',  # Parent category itself
            'pos_terminal_setup',  # Only exists under Phase 2
            'pos_receivables',  # Only exists under Phase 2
            'pos_advanced_settlement',  # Only exists under Phase 2
            'pos_advanced_receivables',  # Only exists under Phase 2
            'pos_delivery',  # Only exists under Phase 2
            'pos_advanced_delivery',  # Only exists under Phase 2
            'pos_advanced_terminal',  # Only exists under Phase 2
            'pos_advanced_day_end',  # Only exists under Phase 2 (different from pos_day_end)
            'pos_code_master',  # Only exists under Phase 2
            'pos_sessions',  # Only exists under Phase 2
            # Note: 'pos_terminal' can be legitimate (main POS) or Phase 2 - keeping it visible
            # Note: 'code_settings' can be used elsewhere - keeping it visible for now
            # Note: 'pos_reports' exists in Reports category too - keeping it visible
        }
        
        # Get menu item details from MenuItemType, filtering out hidden items
        menu_items_detail = []
        for menu_item_id, perms in template.get('menu_items', {}).items():
            # Skip hidden menu items
            if menu_item_id in HIDDEN_MENU_ITEMS:
                continue
                
            try:
                menu_item = MenuItemType.objects.get(menu_item_id=menu_item_id)
                menu_items_detail.append({
                    'menu_item_id': menu_item_id,
                    'display_name': menu_item.display_name,
                    'category': menu_item.category,
                    'subcategory': getattr(menu_item, 'subcategory', None) or (menu_item.transaction_subtype if menu_item.transaction_subtype else None),
                    'menu_type': menu_item.menu_type,
                    'path': menu_item.path,
                    'can_access': perms.get('can_access', False),
                    'can_view': perms.get('can_view', False),
                    'can_create': perms.get('can_create', False),
                    'can_edit': perms.get('can_edit', False),
                    'can_delete': perms.get('can_delete', False),
                })
            except MenuItemType.DoesNotExist:
                # Only include items not in database if they're not hidden
                # (skip hidden items that might exist in templates but not in DB)
                if menu_item_id not in HIDDEN_MENU_ITEMS:
                    menu_items_detail.append({
                        'menu_item_id': menu_item_id,
                        'display_name': menu_item_id.replace('_', ' ').title(),
                        'category': 'Unknown',
                        'menu_type': 'UNKNOWN',
                        'path': '',
                        'can_access': perms.get('can_access', False),
                        'can_view': perms.get('can_view', False),
                        'can_create': perms.get('can_create', False),
                        'can_edit': perms.get('can_edit', False),
                        'can_delete': perms.get('can_delete', False),
                    })
        
        return Response({
            'role': role_template,
            'display_name': template.get('display_name', role_template),
            'description': template.get('description', ''),
            'menu_items': menu_items_detail,  # Array format for display
            'menu_items_dict': template.get('menu_items', {}),  # Raw dictionary format for matrix loading
            'total_items': len(menu_items_detail),
        }, status=status.HTTP_200_OK)


class GetRolePermissionsView(APIView):
    """
    Get all permissions for a specific role.
    Returns permissions in format suitable for frontend matrix.
    Falls back to template permissions if no DB records exist.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, role_key):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can view role permissions'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get all permissions for role (by role_key) from database
        try:
            role_permissions = GroupPermission.objects.filter(role_key=role_key).select_related('menu_item')
            
            # Format for frontend
            permissions_dict = {}
            for perm in role_permissions:
                # Skip if menu_item is None (orphaned permission)
                if not perm.menu_item:
                    continue
                
                permissions_dict[perm.menu_item.menu_item_id] = {
                    'can_access': perm.can_access,
                    'can_view': perm.can_view,
                    'can_create': perm.can_create,
                    'can_edit': perm.can_edit,
                    'can_delete': perm.can_delete,
                }
            
            # If no DB permissions found, fallback to template permissions
            if not permissions_dict:
                template = get_role_permissions(role_key)
                if template and template.get('menu_items'):
                    # Convert template permissions to same format
                    for menu_item_id, perms in template['menu_items'].items():
                        permissions_dict[menu_item_id] = {
                            'can_access': perms.get('can_access', False),
                            'can_view': perms.get('can_view', False),
                            'can_create': perms.get('can_create', False),
                            'can_edit': perms.get('can_edit', False),
                            'can_delete': perms.get('can_delete', False),
                        }
            
            return Response(permissions_dict, status=status.HTTP_200_OK)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error fetching role permissions for role {role_key}: {str(e)}")
            return Response(
                {'error': f'Error fetching permissions: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class BulkRolePermissionView(APIView):
    """
    Bulk save role permissions.
    Creates or updates GroupPermission records for roles.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can manage role permissions'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        from .serializers import BulkRolePermissionSerializer
        serializer = BulkRolePermissionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        permissions_data = serializer.validated_data['permissions']
        created_count = 0
        updated_count = 0
        
        for role_perm_data in permissions_data:
            role_key = role_perm_data['role_key']
            role_permissions = role_perm_data['permissions']
            
            # Get or create Django Group for this role
            group, _ = Group.objects.get_or_create(name=role_key)
            
            for menu_item_id, perms in role_permissions.items():
                try:
                    menu_item = MenuItemType.objects.get(menu_item_id=menu_item_id)
                except MenuItemType.DoesNotExist:
                    continue
                
                permission, created = GroupPermission.objects.update_or_create(
                    group=group,
                    role_key=role_key,
                    menu_item=menu_item,
                    defaults={
                        'can_access': perms.get('can_access', False),
                        'can_view': perms.get('can_view', False),
                        'can_create': perms.get('can_create', False),
                        'can_edit': perms.get('can_edit', False),
                        'can_delete': perms.get('can_delete', False),
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
        
        return Response({
            'message': 'Role permissions updated successfully',
            'created': created_count,
            'updated': updated_count,
        }, status=status.HTTP_200_OK)


class ApplyRoleTemplateToRoleView(APIView):
    """
    Apply a role template to a role (updates GroupPermission records).
    This initializes role permissions from predefined templates.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can apply role templates'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        role_key = request.data.get('role_key')
        template_key = request.data.get('template_key')  # Optional: defaults to role_key if not provided
        
        if not role_key:
            return Response(
                {'error': 'role_key is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Use template_key if provided, otherwise use role_key (role templates match role keys)
        template_to_apply = template_key or role_key
        
        # Get template permissions from role_permissions.py
        from .role_permissions import get_role_permissions
        template_permissions = get_role_permissions(template_to_apply)
        
        if not template_permissions:
            return Response(
                {'error': f'Template not found for role: {template_to_apply}'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get or create Django Group for this role
        group, _ = Group.objects.get_or_create(name=role_key)
        
        created_count = 0
        updated_count = 0
        
        for menu_item_id, perms in template_permissions.items():
            try:
                menu_item = MenuItemType.objects.get(menu_item_id=menu_item_id)
            except MenuItemType.DoesNotExist:
                # Skip hidden menu items that don't exist in database
                continue
            
            permission, created = GroupPermission.objects.update_or_create(
                group=group,
                role_key=role_key,
                menu_item=menu_item,
                defaults={
                    'can_access': perms.get('can_access', False),
                    'can_view': perms.get('can_view', False),
                    'can_create': perms.get('can_create', False),
                    'can_edit': perms.get('can_edit', False),
                    'can_delete': perms.get('can_delete', False),
                }
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        return Response({
            'message': f'Role template "{template_to_apply}" applied successfully to role "{role_key}"',
            'permissions_created': created_count,
            'permissions_updated': updated_count,
        }, status=status.HTTP_200_OK)


class ExportRolePermissionsExcelView(APIView):
    """
    Export role permissions to Excel template.
    Returns an Excel file with all roles and menu items with permissions.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can export role permissions'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get all menu items from database
        menu_items = MenuItemType.objects.all().order_by('category', 'display_name')
        
        # Get all role keys
        role_keys = ['admin', 'posmanager', 'posuser', 'backofficemanager', 'backofficeuser']
        role_labels = {
            'admin': 'Administrator',
            'posmanager': 'POS Manager',
            'posuser': 'POS User',
            'backofficemanager': 'Back Office Manager',
            'backofficeuser': 'Back Office User',
        }
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Role Permissions"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        role_header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create header row
        headers = ['Menu Item', 'Category', 'Subcategory']
        permission_cols = ['Access', 'View', 'Create', 'Edit', 'Delete']
        
        # Add role columns (5 permissions per role)
        for role_key in role_keys:
            for perm in permission_cols:
                headers.append(f"{role_labels[role_key]} - {perm}")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        
        # Get permissions for all roles
        role_permissions = {}
        for role_key in role_keys:
            try:
                perms = GroupPermission.objects.filter(role_key=role_key).select_related('menu_item')
                role_permissions[role_key] = {}
                for perm in perms:
                    if perm.menu_item:
                        role_permissions[role_key][perm.menu_item.menu_item_id] = {
                            'can_access': perm.can_access,
                            'can_view': perm.can_view,
                            'can_create': perm.can_create,
                            'can_edit': perm.can_edit,
                            'can_delete': perm.can_delete,
                        }
            except Exception as e:
                role_permissions[role_key] = {}
        
        # Write data rows
        row_num = 2
        for menu_item in menu_items:
            ws.cell(row=row_num, column=1, value=menu_item.display_name).border = border
            ws.cell(row=row_num, column=2, value=menu_item.category or '').border = border
            ws.cell(row=row_num, column=3, value=(menu_item.transaction_subtype or '')).border = border
            
            col_num = 4
            for role_key in role_keys:
                perms = role_permissions[role_key].get(menu_item.menu_item_id, {})
                for perm_key in ['can_access', 'can_view', 'can_create', 'can_edit', 'can_delete']:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = '✓' if perms.get(perm_key, False) else ''
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                    col_num += 1
            
            row_num += 1
        
        # Freeze panes (first 3 columns)
        ws.freeze_panes = 'D2'
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="role_permissions_template.xlsx"'
        
        wb.save(response)
        return response


class ImportRolePermissionsExcelView(APIView):
    """
    Import role permissions from Excel file.
    Expects an Excel file with the same format as exported template.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can import role permissions'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file uploaded'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from openpyxl import load_workbook
            
            excel_file = request.FILES['file']
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Parse headers to understand role structure
            headers = [cell.value for cell in ws[1]]
            
            # Extract role-permission mapping from headers
            role_permission_map = {}
            role_keys = ['admin', 'posmanager', 'posuser', 'backofficemanager', 'backofficeuser']
            role_labels = {
                'Administrator': 'admin',
                'POS Manager': 'posmanager',
                'POS User': 'posuser',
                'Back Office Manager': 'backofficemanager',
                'Back Office User': 'backofficeuser',
            }
            permission_keys = ['can_access', 'can_view', 'can_create', 'can_edit', 'can_delete']
            permission_labels = ['Access', 'View', 'Create', 'Edit', 'Delete']
            
            # Find column indices for each role-permission combination
            for col_idx, header in enumerate(headers, 1):
                if header and isinstance(header, str) and ' - ' in header:
                    parts = header.split(' - ')
                    if len(parts) == 2:
                        role_label, perm_label = parts
                        role_key = role_labels.get(role_label.strip())
                        perm_idx = permission_labels.index(perm_label.strip()) if perm_label.strip() in permission_labels else -1
                        
                        if role_key and perm_idx >= 0:
                            perm_key = permission_keys[perm_idx]
                            if role_key not in role_permission_map:
                                role_permission_map[role_key] = {}
                            role_permission_map[role_key][perm_key] = col_idx
            
            # Process data rows
            imported_count = 0
            updated_count = 0
            
            for row_idx in range(2, ws.max_row + 1):
                menu_item_name = ws.cell(row=row_idx, column=1).value
                if not menu_item_name:
                    continue
                
                # Find menu item by display_name
                try:
                    menu_item = MenuItemType.objects.get(display_name=menu_item_name)
                except MenuItemType.DoesNotExist:
                    continue
                
                # Process each role
                for role_key, perm_map in role_permission_map.items():
                    perms = {}
                    for perm_key, col_idx in perm_map.items():
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        perms[perm_key] = (cell_value == '✓' or cell_value == 'YES' or cell_value == True or cell_value == 1)
                    
                    # Get or create Django Group
                    group, _ = Group.objects.get_or_create(name=role_key)
                    
                    # Create or update GroupPermission
                    permission, created = GroupPermission.objects.update_or_create(
                        group=group,
                        role_key=role_key,
                        menu_item=menu_item,
                        defaults=perms
                    )
                    
                    if created:
                        imported_count += 1
                    else:
                        updated_count += 1
            
            return Response({
                'message': 'Role permissions imported successfully',
                'imported': imported_count,
                'updated': updated_count,
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Error importing Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


# POS Function Management Viewsets
# Define ViewSets only if models and serializers are available
POSFunctionViewSet = None
RolePOSFunctionMappingViewSet = None

if POSFunction and POSFunctionSerializer:
    class POSFunctionViewSet(ModelViewSet):
        """ViewSet for managing POS functions"""
        queryset = POSFunction.objects.filter(is_active=True)
        serializer_class = POSFunctionSerializer
        permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
        
        def get_queryset(self):
            category = self.request.query_params.get('category', None)
            queryset = POSFunction.objects.filter(is_active=True)
            if category:
                queryset = queryset.filter(category=category)
            return queryset.order_by('category', 'order')


if RolePOSFunctionMapping and RolePOSFunctionMappingSerializer:
    class RolePOSFunctionMappingViewSet(ModelViewSet):
        """ViewSet for managing role-function mappings"""
        queryset = RolePOSFunctionMapping.objects.all()
        serializer_class = RolePOSFunctionMappingSerializer
        permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
        
        def get_queryset(self):
            role = self.request.query_params.get('role', None)
            queryset = RolePOSFunctionMapping.objects.select_related('function', 'created_by')
            if role:
                queryset = queryset.filter(role=role)
            return queryset.order_by('role', 'function__category', 'function__order')
        
        def perform_create(self, serializer):
            """Set created_by to current user"""
            serializer.save(created_by=self.request.user)


class UserPOSPermissionsView(APIView):
    """Get POS function permissions for current user"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        # Admin and superuser have all permissions
        if user.is_superuser or user.is_staff or user.role == 'admin':
            # Return all functions as allowed
            if POSFunction:
                all_functions = POSFunction.objects.filter(is_active=True)
                permissions = {
                    func.function_code: {
                        'allowed': True,
                        'requires_approval': False,
                        'function_name': func.function_name,
                        'keyboard_shortcut': func.keyboard_shortcut,
                        'category': func.category
                    }
                    for func in all_functions
                }
                return Response(permissions)
            return Response({})


class UserAccessibleLocationsView(APIView):
    """
    Get locations accessible to a user based on their role.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, user_id=None):
        # If user_id not provided, use current user
        target_user_id = user_id or request.user.id
        
        # Check if requesting user has permission
        requesting_user = request.user
        try:
            target_user = User.objects.get(id=target_user_id)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Users can see their own accessible locations, admin can see all
        if requesting_user.id != target_user.id and requesting_user.role != 'admin' and not requesting_user.is_superuser:
            return Response(
                {'error': 'Permission denied'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get accessible locations based on role
        accessible_locations = target_user.get_accessible_locations()
        
        # Serialize locations
        locations_data = []
        for location in accessible_locations:
            locations_data.append({
                'id': str(location.id),
                'name': location.name,
                'code': location.code,
                'location_type': location.location_type,
                'is_active': location.is_active,
                'address': location.address,
                'city': location.city,
                'state': location.state,
                'country': location.country,
                'phone': location.phone,
                'email': location.email,
            })
        
        return Response(locations_data)


class UserDefaultLocationView(APIView):
    """
    Get the default location for a user based on their role and mappings.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, user_id=None):
        # If user_id not provided, use current user
        target_user_id = user_id or request.user.id
        
        # Check if requesting user has permission
        requesting_user = request.user
        try:
            target_user = User.objects.get(id=target_user_id)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Users can see their own default location, admin can see all
        if requesting_user.id != target_user.id and requesting_user.role != 'admin' and not requesting_user.is_superuser:
            return Response(
                {'error': 'Permission denied'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get default location
        default_location = target_user.get_default_location()
        
        if default_location:
            return Response({
                'id': str(default_location.id),
                'name': default_location.name,
                'code': default_location.code,
                'location_type': default_location.location_type,
                'is_active': default_location.is_active,
                'address': default_location.address,
                'city': default_location.city,
                'state': default_location.state,
                'country': default_location.country,
                'phone': default_location.phone,
                'email': default_location.email,
            })
        else:
            return Response(None)


class SyncUserLocationMappingsView(APIView):
    """
    Sync user location mappings based on their role.
    This creates or updates mappings according to role-based rules.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, user_id=None):
        # Only admin can sync mappings
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can sync location mappings'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # If user_id not provided, sync all users
        target_user_id = user_id or request.data.get('user_id')
        
        if target_user_id:
            # Sync specific user
            try:
                target_user = User.objects.get(id=target_user_id)
                result = self.sync_user_mappings(target_user)
                return Response(result)
            except User.DoesNotExist:
                return Response(
                    {'error': 'User not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
        else:
            # Sync all users
            users = User.objects.all()
            results = []
            for user in users:
                result = self.sync_user_mappings(user)
                results.append({
                    'user_id': str(user.id),
                    'username': user.username,
                    'role': user.role,
                    'result': result
                })
            
            return Response({
                'message': 'All users synced successfully',
                'results': results
            })
    
    def sync_user_mappings(self, user):
        """Sync location mappings for a single user based on their role."""
        from .models import UserLocationMapping
        from organization.models import Location
        
        # Get accessible locations based on role
        accessible_locations = user.get_accessible_locations()
        
        # Remove existing mappings that are no longer accessible
        UserLocationMapping.objects.filter(user=user).exclude(
            location__in=accessible_locations
        ).delete()
        
        created_count = 0
        updated_count = 0
        
        # Create or update mappings for accessible locations
        for location in accessible_locations:
            # Determine access type based on role and location type
            if user.role == 'posuser':
                access_type = 'pos'
            elif user.role in ['backofficemanager', 'backofficeuser']:
                access_type = 'back_office'
            elif user.role == 'posmanager':
                access_type = 'both'  # POS managers can access both POS and back office
            elif user.role == 'admin':
                access_type = 'both'  # Admins can access both
            else:
                access_type = 'back_office'
            
            # For POS users, only create one mapping (first location)
            if user.role == 'posuser':
                existing_mapping = UserLocationMapping.objects.filter(
                    user=user,
                    access_type__in=['pos', 'both']
                ).first()
                
                if existing_mapping:
                    if existing_mapping.location_id != location.id:
                        existing_mapping.location = location
                        existing_mapping.save()
                        updated_count += 1
                else:
                    UserLocationMapping.objects.create(
                        user=user,
                        location=location,
                        access_type=access_type,
                        is_default=True,
                        created_by=self.request.user
                    )
                    created_count += 1
                break  # Only one location for POS users
            else:
                # For other roles, create mapping for each accessible location
                mapping, created = UserLocationMapping.objects.get_or_create(
                    user=user,
                    location=location,
                    access_type=access_type,
                    defaults={
                        'is_active': True,
                        'is_default': False,
                        'created_by': self.request.user,
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
        
        return {
            'message': f'User {user.username} synced successfully',
            'accessible_locations': accessible_locations.count(),
            'created': created_count,
            'updated': updated_count,
        }


# User Location Mapping Views
class UserLocationMappingViewSet(ModelViewSet):
    """
    ViewSet for managing user-location mappings.
    Only admin role can manage mappings.
    """
    serializer_class = UserLocationMappingSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        # Only admin role can manage mappings
        if user.role != 'admin' and not user.is_superuser:
            return UserLocationMapping.objects.none()
        
        return UserLocationMapping.objects.select_related('user', 'location', 'created_by')
    
    def create(self, request, *args, **kwargs):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can manage user-location mappings'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can manage user-location mappings'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)
    
    def destroy(self, request, *args, **kwargs):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can manage user-location mappings'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)


class BulkUserLocationMappingView(APIView):
    """
    Bulk create/update user-location mappings.
    Only admin role can perform bulk operations.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        # Check admin permission
        if request.user.role != 'admin' and not request.user.is_superuser:
            return Response(
                {'error': 'Only admin can manage user-location mappings'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = BulkUserLocationMappingSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        mappings_data = serializer.validated_data['mappings']
        created_count = 0
        updated_count = 0
        
        # Import UserLocationMapping model
        from .models import UserLocationMapping
        
        for mapping_data in mappings_data:
            try:
                user = User.objects.get(id=mapping_data['user_id'])
                location = None  # Will be set below
                access_type = mapping_data['access_type']
                is_active = mapping_data.get('is_active', True)
                is_default = mapping_data.get('is_default', False)
                
                # Get location by ID
                from organization.models import Location
                location = Location.objects.get(id=mapping_data['location_id'])
                
                # Create or update mapping
                mapping, created = UserLocationMapping.objects.update_or_create(
                    user=user,
                    location=location,
                    access_type=access_type,
                    defaults={
                        'is_active': is_active,
                        'is_default': is_default,
                        'created_by': request.user,
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    # Update existing mapping
                    mapping.is_active = is_active
                    mapping.is_default = is_default
                    mapping.save()
                    updated_count += 1
                    
            except (User.DoesNotExist, Location.DoesNotExist) as e:
                continue  # Skip invalid mappings
        
        return Response({
            'message': 'User-location mappings updated successfully',
            'created': created_count,
            'updated': updated_count,
        }, status=status.HTTP_200_OK)
        
        # Get role-based permissions
        role = user.role
        
        if not role or role not in ['posmanager', 'posuser']:
            # Not a POS role, return empty permissions
            return Response({})
        
        if RolePOSFunctionMapping:
            # Get all allowed functions for this role
            mappings = RolePOSFunctionMapping.objects.filter(
                role=role,
                is_allowed=True
            ).select_related('function')
            
            permissions = {
                mapping.function.function_code: {
                    'allowed': True,
                    'requires_approval': mapping.requires_approval,
                    'function_name': mapping.function.function_name,
                    'keyboard_shortcut': mapping.function.keyboard_shortcut,
                    'category': mapping.function.category
                }
                for mapping in mappings
            }
            
            return Response(permissions)
        
        return Response({})
