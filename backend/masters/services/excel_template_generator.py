import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from django.apps import apps
from django.db import models
from django.contrib.contenttypes.models import ContentType
import uuid
from datetime import datetime
import os


class ExcelTemplateGenerator:
    """
    Generate Excel templates for master data upload
    """
    
    def __init__(self):
        self.workbook = None
        self.styles = {}
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup styles for the Excel template"""
        # Header style
        self.styles['header'] = Font(
            name='Calibri',
            size=12,
            bold=True,
            color='FFFFFF'
        )
        self.styles['header_fill'] = PatternFill(
            start_color='366092',
            end_color='366092',
            fill_type='solid'
        )
        
        # Title style
        self.styles['title'] = Font(
            name='Calibri',
            size=16,
            bold=True,
            color='366092'
        )
        
        # Instruction style
        self.styles['instruction'] = Font(
            name='Calibri',
            size=11,
            color='333333'
        )
        
        # Hyperlink style
        self.styles['hyperlink'] = Font(
            name='Calibri',
            size=11,
            color='0563C1',
            underline='single'
        )
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.styles['border'] = thin_border
        
        # Alignment
        self.styles['center_alignment'] = Alignment(
            horizontal='center',
            vertical='center'
        )
        self.styles['left_alignment'] = Alignment(
            horizontal='left',
            vertical='center'
        )
    
    def generate_master_data_template(self, include_sample_data=False):
        """
        Generate the complete master data template with all sheets
        """
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create all sheets
        self._create_instructions_sheet()
        self._create_organization_setup_sheet(include_sample_data)
        self._create_general_masters_sheet(include_sample_data)
        self._create_item_data_sheet(include_sample_data)
        self._create_attributes_sheet(include_sample_data)
        self._create_attribute_values_sheet(include_sample_data)
        self._create_customers_sheet(include_sample_data)
        self._create_vendors_sheet(include_sample_data)
        self._create_tax_configuration_sheet(include_sample_data)
        self._create_additional_masters_sheet(include_sample_data)
        
        return self.workbook
    
    def _create_instructions_sheet(self):
        """
        Create the instructions sheet with hyperlinks to all other sheets
        """
        ws = self.workbook.create_sheet('Instructions & Navigation')
        
        # Title
        ws['A1'] = 'Master Data Upload Template'
        ws['A1'].font = self.styles['title']
        ws.merge_cells('A1:B1')
        
        # Generation info
        ws['A3'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A3'].font = self.styles['instruction']
        
        ws['A4'] = 'Version: 1.0'
        ws['A4'].font = self.styles['instruction']
        
        # Instructions
        ws['A6'] = 'Instructions:'
        ws['A6'].font = Font(name='Calibri', size=12, bold=True)
        
        instructions = [
            '1. This Excel template contains multiple sheets for different master data types.',
            '2. Click on the sheet names below to navigate to specific master data sheets.',
            '3. Fill in the required data in each sheet (marked with * are required fields).',
            '4. Follow the data format and validation rules specified in each sheet.',
            '5. Save the completed template and upload it through the Django Admin interface.',
            '6. The system will validate the data and provide detailed error reports if needed.',
            '',
            'Important Notes:',
            '- Do not change the column headers or sheet names.',
            '- Do not delete or reorder columns.',
            '- Required fields are marked with an asterisk (*).',
            '- Follow the specified data formats for each field.',
            '- Check for duplicate records before uploading.',
            '- Large files may take longer to process.',
            '',
            'Navigate to Master Data Sheets:'
        ]
        
        row = 8
        for instruction in instructions:
            ws[f'A{row}'] = instruction
            ws[f'A{row}'].font = self.styles['instruction']
            ws[f'A{row}'].alignment = self.styles['left_alignment']
            row += 1
        
        # Hyperlinks to sheets
        ws[f'A{row}'] = 'Master Data Sheets:'
        ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
        row += 2
        
        sheet_links = [
            ('Organization Setup', 'Company and Location information'),
            ('General Masters', 'Categories, UOM, Payment Modes'),
            ('Item Data', 'Products and Item Master information'),
            ('Attributes', 'Product attributes and definitions'),
            ('Attribute Values', 'Attribute value mappings'),
            ('Customers', 'Customer information and details'),
            ('Vendors', 'Vendor/Supplier information'),
            ('Tax Configuration', 'Tax codes and rates'),
            ('Additional Masters', 'Brands, Departments, Divisions')
        ]
        
        for sheet_name, description in sheet_links:
            # Create hyperlink
            ws[f'A{row}'] = sheet_name
            ws[f'A{row}'].font = self.styles['hyperlink']
            ws[f'A{row}'].hyperlink = Hyperlink(
                ref=f'A{row}',
                location=f"'{sheet_name}'!A1",
                tooltip=f'Go to {sheet_name} sheet'
            )
            
            ws[f'B{row}'] = f"- {description}"
            ws[f'B{row}'].font = self.styles['instruction']
            ws[f'B{row}'].alignment = self.styles['left_alignment']
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        
        # Add some styling to the header row
        for col in ['A', 'B']:
            cell = ws[f'{col}1']
            cell.fill = self.styles['header_fill']
            cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            cell.alignment = self.styles['center_alignment']
    
    def _create_organization_setup_sheet(self, include_sample_data=False):
        """Create Organization Setup sheet"""
        ws = self.workbook.create_sheet('Organization Setup')
        
        # Company section
        ws['A1'] = 'Company Information'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells('A1:D1')
        
        company_headers = ['Field Name', 'Required', 'Description', 'Sample Data']
        company_fields = [
            ('name*', 'Yes', 'Company name', 'ABC Retail Store'),
            ('code*', 'Yes', 'Unique company code (uppercase letters and numbers only)', 'ABC001'),
            ('description', 'No', 'Company description', 'Leading retail chain in electronics'),
            ('address', 'No', 'Complete address', '123 Main Street, City, State'),
            ('city', 'No', 'City', 'New York'),
            ('state', 'No', 'State/Province', 'NY'),
            ('country', 'No', 'Country', 'United States'),
            ('postal_code', 'No', 'Postal/ZIP code', '10001'),
            ('phone', 'No', 'Phone number', '+1-212-555-1234'),
            ('email', 'No', 'Email address', 'info@abcretail.com'),
            ('website', 'No', 'Website URL', 'https://www.abcretail.com'),
            ('tax_id', 'No', 'Tax identification number', '12-3456789'),
            ('registration_number', 'No', 'Business registration number', 'REG123456'),
            ('currency*', 'Yes', 'Currency code (USD, EUR, INR, etc.)', 'USD'),
            ('timezone*', 'Yes', 'Timezone', 'America/New_York')
        ]
        
        self._write_sheet_section(ws, company_headers, company_fields, 2, include_sample_data)
        
        # Location section
        current_row = len(company_fields) + 4
        ws[f'A{current_row}'] = 'Location Information'
        ws[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells(f'A{current_row}:D{current_row}')
        
        location_headers = ['Field Name', 'Required', 'Description', 'Sample Data']
        location_fields = [
            ('name*', 'Yes', 'Location name', 'Manhattan Store'),
            ('code*', 'Yes', 'Unique location code (uppercase letters and numbers only)', 'NYC001'),
            ('description', 'No', 'Location description', 'Flagship store in Manhattan'),
            ('company_code*', 'Yes', 'Company code this location belongs to', 'ABC001'),
            ('address', 'No', 'Complete address', '456 Broadway, New York'),
            ('city', 'No', 'City', 'New York'),
            ('state', 'No', 'State/Province', 'NY'),
            ('country', 'No', 'Country', 'United States'),
            ('postal_code', 'No', 'Postal/ZIP code', '10013'),
            ('phone', 'No', 'Phone number', '+1-212-555-5678'),
            ('email', 'No', 'Email address', 'manhattan@abcretail.com'),
            ('manager', 'No', 'Location manager name', 'John Smith'),
            ('location_type*', 'Yes', 'Type: store, headquarters, warehouse, distribution, factory, showroom', 'store'),
            ('latitude', 'No', 'GPS latitude coordinate', '40.7128'),
            ('longitude', 'No', 'GPS longitude coordinate', '-74.0060')
        ]
        
        self._write_sheet_section(ws, location_headers, location_fields, current_row + 1, include_sample_data)
    
    def _create_general_masters_sheet(self, include_sample_data=False):
        """Create General Masters sheet"""
        ws = self.workbook.create_sheet('General Masters')
        
        ws['A1'] = 'General Masters - Categories, UOM, Payment Modes'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells('A1:E1')
        
        headers = ['Master Type', 'Field Name', 'Required', 'Description', 'Sample Data']
        
        # Categories
        category_fields = [
            ('Category', 'name*', 'Yes', 'Category name', 'Electronics'),
            ('Category', 'description', 'No', 'Category description', 'Electronic devices and accessories'),
            ('Category', 'parent_name', 'No', 'Parent category name (for sub-categories)', 'Computers'),
            ('Category', 'is_active', 'No', 'Active status (True/False)', 'True'),
            ('Category', 'sort_order', 'No', 'Display order (number)', '1')
        ]
        
        # UOM
        uom_fields = [
            ('UOM', 'code*', 'Yes', 'Unit of Measure code (uppercase)', 'PCS'),
            ('UOM', 'description*', 'Yes', 'Description', 'Pieces'),
            ('UOM', 'basis*', 'Yes', 'Basis: length, units, volume, capacity', 'units'),
            ('UOM', 'decimals', 'No', 'Decimal places (number)', '0'),
            ('UOM', 'is_stock_uom', 'No', 'Used for stock (True/False)', 'True'),
            ('UOM', 'is_purchase_uom', 'No', 'Used for purchase (True/False)', 'True'),
            ('UOM', 'is_sales_uom', 'No', 'Used for sales (True/False)', 'True')
        ]
        
        # Payment Modes
        payment_fields = [
            ('Payment Mode', 'name*', 'Yes', 'Payment mode name', 'Cash'),
            ('Payment Mode', 'code*', 'Yes', 'Payment mode code', 'CASH'),
            ('Payment Mode', 'description', 'No', 'Description', 'Cash payment'),
            ('Payment Mode', 'is_active', 'No', 'Active status (True/False)', 'True')
        ]
        
        # Combine all fields
        all_fields = category_fields + uom_fields + payment_fields
        self._write_sheet_section(ws, headers, all_fields, 2, include_sample_data)
    
    def _create_item_data_sheet(self, include_sample_data=False):
        """Create Item Data sheet"""
        ws = self.workbook.create_sheet('Item Data')
        
        ws['A1'] = 'Item Data - Products and Item Master'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells('A1:F1')
        
        headers = ['Item Type', 'Field Name', 'Required', 'Description', 'Sample Data', 'Notes']
        
        # Basic Product
        product_fields = [
            ('Product', 'name*', 'Yes', 'Product name', 'Laptop Computer'),
            ('Product', 'sku*', 'Yes', 'Unique SKU', 'LAPTOP-001'),
            ('Product', 'barcode', 'No', 'Barcode', '1234567890123'),
            ('Product', 'description', 'No', 'Product description', 'High-performance laptop'),
            ('Product', 'price*', 'Yes', 'Selling price', '999.99'),
            ('Product', 'cost', 'No', 'Cost price', '750.00'),
            ('Product', 'stock_quantity', 'No', 'Current stock', '50'),
            ('Product', 'minimum_stock', 'No', 'Minimum stock level', '10'),
            ('Product', 'category_name', 'No', 'Category name', 'Electronics'),
            ('Product', 'is_active', 'No', 'Active status (True/False)', 'True'),
            ('Product', 'is_taxable', 'No', 'Taxable (True/False)', 'True'),
            ('Product', 'tax_rate', 'No', 'Tax rate percentage', '8.25')
        ]
        
        # Advanced Item Master
        item_fields = [
            ('Item Master', 'item_code*', 'Yes', 'Unique item code', 'ITEM-001'),
            ('Item Master', 'item_name*', 'Yes', 'Item name', 'Gaming Laptop'),
            ('Item Master', 'short_name', 'No', 'Short name', 'Gaming Laptop'),
            ('Item Master', 'brand', 'No', 'Brand name', 'TechBrand'),
            ('Item Master', 'supplier', 'No', 'Supplier name', 'TechSupplier'),
            ('Item Master', 'cost_price*', 'Yes', 'Cost price', '750.00'),
            ('Item Master', 'sell_price*', 'Yes', 'Selling price', '999.99'),
            ('Item Master', 'mrp', 'No', 'Maximum retail price', '1099.99'),
            ('Item Master', 'category_name', 'No', 'Category name', 'Electronics'),
            ('Item Master', 'tax_code', 'No', 'Tax code', 'GST-18'),
            ('Item Master', 'hsn_code', 'No', 'HSN code', '847130'),
            ('Item Master', 'material_type', 'No', 'Material type: raw, finished, semi, consumable, service', 'finished'),
            ('Item Master', 'item_type', 'No', 'Item type: spare, device, ew, accessory', 'device'),
            ('Item Master', 'base_uom_code', 'No', 'Base UOM code', 'PCS'),
            ('Item Master', 'is_active', 'No', 'Active status (True/False)', 'True')
        ]
        
        # Combine all fields
        all_fields = product_fields + item_fields
        self._write_sheet_section(ws, headers, all_fields, 2, include_sample_data)
    
    def _create_attributes_sheet(self, include_sample_data=False):
        """Create Attributes sheet"""
        ws = self.workbook.create_sheet('Attributes')
        
        ws['A1'] = 'Product Attributes'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells('A1:E1')
        
        headers = ['Field Name', 'Required', 'Description', 'Sample Data', 'Data Type Options']
        
        fields = [
            ('name*', 'Yes', 'Attribute name', 'Color', 'text, number, boolean, date, select'),
            ('description', 'No', 'Attribute description', 'Product color variation', ''),
            ('data_type*', 'Yes', 'Data type', 'select', 'text, number, boolean, date, select'),
            ('is_active', 'No', 'Active status (True/False)', 'True', ''),
            ('sort_order', 'No', 'Display order', '1', '')
        ]
        
        self._write_sheet_section(ws, headers, fields, 2, include_sample_data)
    
    def _create_attribute_values_sheet(self, include_sample_data=False):
        """Create Attribute Values sheet"""
        ws = self.workbook.create_sheet('Attribute Values')
        
        ws['A1'] = 'Attribute Values'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells('A1:E1')
        
        headers = ['Field Name', 'Required', 'Description', 'Sample Data', 'Notes']
        
        fields = [
            ('attribute_name*', 'Yes', 'Attribute name this value belongs to', 'Color', 'Must match an attribute name'),
            ('value*', 'Yes', 'Attribute value', 'Red', 'Actual value for the attribute'),
            ('description', 'No', 'Value description', 'Bright red color', ''),
            ('is_active', 'No', 'Active status (True/False)', 'True', ''),
            ('sort_order', 'No', 'Display order', '1', '')
        ]
        
        self._write_sheet_section(ws, headers, fields, 2, include_sample_data)
    
    def _create_customers_sheet(self, include_sample_data=False):
        """Create Customers sheet"""
        ws = self.workbook.create_sheet('Customers')
        
        ws['A1'] = 'Customer Information'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells('A1:E1')
        
        headers = ['Field Name', 'Required', 'Description', 'Sample Data', 'Notes']
        
        fields = [
            ('first_name*', 'Yes', 'Customer first name', 'John', ''),
            ('last_name*', 'Yes', 'Customer last name', 'Doe', ''),
            ('company_name', 'No', 'Company name (for business customers)', 'ABC Corp', 'Required for business customers'),
            ('customer_type*', 'Yes', 'Customer type: individual, business, wholesale, vip', 'individual', ''),
            ('email', 'No', 'Email address', 'john.doe@email.com', ''),
            ('phone*', 'Yes', 'Phone number', '+1-212-555-1234', 'Include country code'),
            ('mobile', 'No', 'Mobile number', '+1-212-555-5678', ''),
            ('address_line_1', 'No', 'Street address', '123 Main St', ''),
            ('city', 'No', 'City', 'New York', ''),
            ('state', 'No', 'State/Province', 'NY', ''),
            ('postal_code', 'No', 'Postal/ZIP code', '10001', ''),
            ('country', 'No', 'Country', 'United States', ''),
            ('credit_limit', 'No', 'Credit limit', '5000.00', 'For credit customers'),
            ('discount_percentage', 'No', 'Default discount percentage', '5.00', ''),
            ('is_active', 'No', 'Active status (True/False)', 'True', ''),
            ('is_vip', 'No', 'VIP status (True/False)', 'False', ''),
            ('allow_credit', 'No', 'Allow credit purchases (True/False)', 'False', ''),
            ('notes', 'No', 'Additional notes', 'Preferred customer', '')
        ]
        
        self._write_sheet_section(ws, headers, fields, 2, include_sample_data)
    
    def _create_vendors_sheet(self, include_sample_data=False):
        """Create Vendors sheet"""
        ws = self.workbook.create_sheet('Vendors')
        
        ws['A1'] = 'Vendor/Supplier Information'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells('A1:E1')
        
        headers = ['Field Name', 'Required', 'Description', 'Sample Data', 'Notes']
        
        fields = [
            ('name*', 'Yes', 'Vendor/Supplier name', 'TechSupplier Inc', ''),
            ('code*', 'Yes', 'Unique vendor code', 'VENDOR-001', 'Uppercase letters and numbers'),
            ('description', 'No', 'Vendor description', 'Technology equipment supplier', ''),
            ('contact_person', 'No', 'Contact person name', 'Jane Smith', ''),
            ('email', 'No', 'Email address', 'contact@techsupplier.com', ''),
            ('phone*', 'Yes', 'Phone number', '+1-212-555-9876', 'Include country code'),
            ('mobile', 'No', 'Mobile number', '+1-212-555-5432', ''),
            ('address', 'No', 'Complete address', '789 Supplier Ave, Tech City', ''),
            ('city', 'No', 'City', 'San Francisco', ''),
            ('state', 'No', 'State/Province', 'CA', ''),
            ('postal_code', 'No', 'Postal/ZIP code', '94105', ''),
            ('country', 'No', 'Country', 'United States', ''),
            ('website', 'No', 'Website URL', 'https://www.techsupplier.com', ''),
            ('tax_id', 'No', 'Tax identification number', '98-7654321', ''),
            ('payment_terms', 'No', 'Payment terms', 'NET 30', ''),
            ('is_active', 'No', 'Active status (True/False)', 'True', ''),
            ('notes', 'No', 'Additional notes', 'Primary electronics supplier', '')
        ]
        
        self._write_sheet_section(ws, headers, fields, 2, include_sample_data)
    
    def _create_tax_configuration_sheet(self, include_sample_data=False):
        """Create Tax Configuration sheet"""
        ws = self.workbook.create_sheet('Tax Configuration')
        
        ws['A1'] = 'Tax Configuration'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells('A1:E1')
        
        headers = ['Field Name', 'Required', 'Description', 'Sample Data', 'Notes']
        
        fields = [
            ('name*', 'Yes', 'Tax name', 'GST', ''),
            ('code*', 'Yes', 'Tax code', 'GST-18', 'Unique identifier'),
            ('description', 'No', 'Tax description', 'Goods and Services Tax', ''),
            ('rate*', 'Yes', 'Tax rate percentage', '18.00', 'Percentage value'),
            ('tax_type*', 'Yes', 'Tax type: sales, purchase, both', 'sales', ''),
            ('is_active', 'No', 'Active status (True/False)', 'True', ''),
            ('effective_from', 'No', 'Effective from date', '2024-01-01', 'YYYY-MM-DD format'),
            ('effective_to', 'No', 'Effective to date', '', 'YYYY-MM-DD format, leave blank if ongoing'),
            ('applicable_states', 'No', 'Applicable states', 'CA,NY,TX', 'Comma separated state codes'),
            ('notes', 'No', 'Additional notes', 'Standard GST rate', '')
        ]
        
        self._write_sheet_section(ws, headers, fields, 2, include_sample_data)
    
    def _create_additional_masters_sheet(self, include_sample_data=False):
        """Create Additional Masters sheet"""
        ws = self.workbook.create_sheet('Additional Masters')
        
        ws['A1'] = 'Additional Masters - Brands, Departments, Divisions'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='366092')
        ws.merge_cells('A1:E1')
        
        headers = ['Master Type', 'Field Name', 'Required', 'Description', 'Sample Data']
        
        # Brands
        brand_fields = [
            ('Brand', 'name*', 'Yes', 'Brand name', 'TechBrand'),
            ('Brand', 'code', 'No', 'Brand code', 'TB001', 'Unique identifier'),
            ('Brand', 'description', 'No', 'Brand description', 'Leading technology brand', ''),
            ('Brand', 'is_active', 'No', 'Active status (True/False)', 'True', '')
        ]
        
        # Departments
        department_fields = [
            ('Department', 'name*', 'Yes', 'Department name', 'Electronics'),
            ('Department', 'code', 'No', 'Department code', 'ELEC001', 'Unique identifier'),
            ('Department', 'description', 'No', 'Department description', 'Electronic devices and accessories', ''),
            ('Department', 'is_active', 'No', 'Active status (True/False)', 'True', '')
        ]
        
        # Divisions
        division_fields = [
            ('Division', 'name*', 'Yes', 'Division name', 'Consumer Electronics'),
            ('Division', 'code', 'No', 'Division code', 'CE001', 'Unique identifier'),
            ('Division', 'description', 'No', 'Division description', 'Consumer electronics division', ''),
            ('Division', 'is_active', 'No', 'Active status (True/False)', 'True', '')
        ]
        
        # Combine all fields
        all_fields = brand_fields + department_fields + division_fields
        self._write_sheet_section(ws, headers, all_fields, 2, include_sample_data)
    
    def _write_sheet_section(self, worksheet, headers, fields, start_row, include_sample_data=False):
        """
        Write a section to the worksheet with headers and fields
        """
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center_alignment']
            cell.border = self.styles['border']
        
        # Write field data
        for row_idx, field_data in enumerate(fields, start_row + 1):
            for col_idx, value in enumerate(field_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles['instruction']
                cell.alignment = self.styles['left_alignment']
                cell.border = self.styles['border']
                
                # Add styling for required fields
                if col_idx == 1 and '*' in str(value):
                    cell.font = Font(name='Calibri', size=11, bold=True, color='FF0000')
        
        # Adjust column widths
        column_widths = [25, 15, 40, 20, 30]
        for col_idx, width in enumerate(column_widths, 1):
            if col_idx <= len(headers):
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Add sample data separator if needed
        if include_sample_data and fields:
            separator_row = start_row + len(fields) + 1
            worksheet.cell(row=separator_row, column=1, value='--- Sample Data Above ---')
            worksheet.cell(row=separator_row, column=1).font = Font(
                name='Calibri', 
                size=10, 
                italic=True, 
                color='666666'
            )
            # Only merge if we have valid columns
            if len(headers) > 0:
                end_col = get_column_letter(len(headers))
                worksheet.merge_cells(f'{separator_row}A:{separator_row}{end_col}')


def generate_template_file(include_sample_data=False):
    """
    Generate the Excel template file and return the workbook
    """
    generator = ExcelTemplateGenerator()
    return generator.generate_master_data_template(include_sample_data)


def save_template_to_file(workbook, file_path):
    """
    Save the workbook to a file
    """
    workbook.save(file_path)
    return file_path
