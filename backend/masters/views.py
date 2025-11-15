from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_protect
import json
import openpyxl
import io
from datetime import datetime

from .models import UploadSession
from .services.excel_template_generator import generate_template_file


@staff_member_required
def download_template_standalone(request):
    """
    Standalone template download view (not tied to Django admin)
    """
    try:
        include_sample_data = request.GET.get('sample_data', 'false').lower() == 'true'
        workbook = generate_template_file(include_sample_data)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Set filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if include_sample_data:
            filename = f'Master_Data_Template_Sample_{timestamp}.xlsx'
        else:
            filename = f'Master_Data_Template_{timestamp}.xlsx'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        virtual_workbook = io.BytesIO()
        workbook.save(virtual_workbook)
        virtual_workbook.seek(0)
        response.write(virtual_workbook.read())
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating template: {str(e)}')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/masters/upload/'))


@staff_member_required
def master_data_upload(request):
    """
    Simplified master data upload view
    """
    context = {
        'title': 'Master Data Upload',
        'opts': UploadSession._meta,
        'has_change_permission': True,
    }
    return render(request, 'masters/upload.html', context)


@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def process_upload(request):
    """
    Process uploaded Excel file synchronously and return JSON response
    """
    try:
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return JsonResponse({
                'success': False,
                'error': 'No file uploaded'
            }, status=400)
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'error': 'Please upload a valid Excel file (.xlsx or .xls)'
            }, status=400)
        
        # Create upload session for tracking
        session = UploadSession.objects.create(
            session_name=f"Upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            excel_file=excel_file,
            file_name=excel_file.name,
            file_size=excel_file.size,
            created_by=request.user,
            status='processing'
        )
        
        # Process the file synchronously
        result = process_excel_file(excel_file, session)
        
        # Update session with results
        session.status = 'completed' if result['success'] else 'failed'
        session.total_records = result['total_records']
        session.successful_records = result['successful_records']
        session.failed_records = result['failed_records']
        session.progress_percentage = 100
        session.save()
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error processing file: {str(e)}'
        }, status=500)


def process_excel_file(excel_file, session):
    """
    Process the uploaded Excel file synchronously
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)
        
        total_records = 0
        successful_records = 0
        failed_records = 0
        error_details = []
        
        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Skip if sheet is empty or has no data
            if sheet.max_row <= 1:
                continue
            
            # Get header row
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # Process data rows
            for row_num in range(2, sheet.max_row + 1):
                total_records += 1
                row_data = {}
                
                for col_num, header in enumerate(headers):
                    if header:
                        cell_value = sheet.cell(row=row_num, column=col_num + 1).value
                        row_data[header] = cell_value
                
                # Validate and process the row
                try:
                    # Basic validation
                    if not any(row_data.values()):
                        failed_records += 1
                        error_details.append({
                            'sheet': sheet_name,
                            'row': row_num,
                            'error': 'Empty row'
                        })
                        continue
                    
                    # Process the data based on sheet type
                    success = process_row_data(sheet_name, row_data)
                    
                    if success:
                        successful_records += 1
                    else:
                        failed_records += 1
                        error_details.append({
                            'sheet': sheet_name,
                            'row': row_num,
                            'error': 'Failed to process record'
                        })
                        
                except Exception as e:
                    failed_records += 1
                    error_details.append({
                        'sheet': sheet_name,
                        'row': row_num,
                        'error': str(e)
                    })
        
        # Create error file if there are errors
        error_file_url = None
        if error_details:
            error_file_url = create_error_file(workbook, error_details, session)
        
        return {
            'success': True,
            'total_records': total_records,
            'successful_records': successful_records,
            'failed_records': failed_records,
            'error_file_url': error_file_url,
            'session_id': str(session.id)
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Error processing Excel file: {str(e)}',
            'total_records': 0,
            'successful_records': 0,
            'failed_records': 0
        }


def process_row_data(sheet_name, row_data):
    """
    Process a single row of data based on the sheet type
    """
    try:
        # This is where you would implement the actual business logic
        # for processing different types of master data
        
        if sheet_name.lower() in ['categories', 'items', 'products']:
            # Process product/item data
            return process_product_data(row_data)
        elif sheet_name.lower() in ['customers']:
            # Process customer data
            return process_customer_data(row_data)
        elif sheet_name.lower() in ['suppliers', 'vendors']:
            # Process supplier data
            return process_supplier_data(row_data)
        elif sheet_name.lower() in ['organization', 'locations']:
            # Process organization/location data
            return process_organization_data(row_data)
        else:
            # Generic processing for other sheets
            return process_generic_data(row_data)
            
    except Exception as e:
        print(f"Error processing row in {sheet_name}: {e}")
        return False


def process_product_data(row_data):
    """
    Process product/item data
    """
    # Implement actual product processing logic here
    # For now, just validate that required fields exist
    required_fields = ['name', 'code']
    for field in required_fields:
        if field not in row_data or not row_data[field]:
            return False
    
    # TODO: Create/update product in database
    # Product.objects.update_or_create(
    #     code=row_data['code'],
    #     defaults={'name': row_data['name'], ...}
    # )
    
    return True


def process_customer_data(row_data):
    """
    Process customer data
    """
    # Implement actual customer processing logic here
    required_fields = ['name', 'code']
    for field in required_fields:
        if field not in row_data or not row_data[field]:
            return False
    
    # TODO: Create/update customer in database
    # Customer.objects.update_or_create(
    #     code=row_data['code'],
    #     defaults={'name': row_data['name'], ...}
    # )
    
    return True


def process_supplier_data(row_data):
    """
    Process supplier data
    """
    # Implement actual supplier processing logic here
    required_fields = ['name', 'code']
    for field in required_fields:
        if field not in row_data or not row_data[field]:
            return False
    
    # TODO: Create/update supplier in database
    # Supplier.objects.update_or_create(
    #     code=row_data['code'],
    #     defaults={'name': row_data['name'], ...}
    # )
    
    return True


def process_organization_data(row_data):
    """
    Process organization/location data
    """
    # Implement actual organization processing logic here
    required_fields = ['name', 'code']
    for field in required_fields:
        if field not in row_data or not row_data[field]:
            return False
    
    # TODO: Create/update organization in database
    # Organization.objects.update_or_create(
    #     code=row_data['code'],
    #     defaults={'name': row_data['name'], ...}
    # )
    
    return True


def process_generic_data(row_data):
    """
    Generic data processing for other sheet types
    """
    # Basic validation - check if row has any data
    return any(row_data.values())


def create_error_file(workbook, error_details, session):
    """
    Create an error file with error details added to the original Excel file
    """
    try:
        # Create a copy of the workbook for errors
        error_workbook = workbook
        
        # Add error details to each sheet
        for sheet_name in error_workbook.sheetnames:
            sheet = error_workbook[sheet_name]
            
            # Add error column header if it doesn't exist
            max_col = sheet.max_column
            error_col = max_col + 1
            sheet.cell(row=1, column=error_col, value="ERROR_STATUS")
            
            # Add error messages to rows
            sheet_errors = [e for e in error_details if e['sheet'] == sheet_name]
            for error in sheet_errors:
                row_num = error['row']
                sheet.cell(row=row_num, column=error_col, value=error['error'])
        
        # Save the error file
        error_filename = f"Master_Data_Upload_Errors_{session.session_name}.xlsx"
        virtual_workbook = io.BytesIO()
        error_workbook.save(virtual_workbook)
        virtual_workbook.seek(0)
        
        # Save to session's file field (or create a temporary file)
        session.error_file.save(error_filename, virtual_workbook)
        
        # Return URL for download
        return reverse('admin:masters_uploadsession_download_error_file', args=[session.id])
        
    except Exception as e:
        print(f"Error creating error file: {e}")
        return None


@staff_member_required
def download_error_file(request, session_id):
    """
    Download the error file for a specific upload session
    """
    try:
        session = UploadSession.objects.get(id=session_id)
        if not session.error_file:
            return JsonResponse({'error': 'No error file available'}, status=404)
        
        # Serve the error file
        response = HttpResponse(
            session.error_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{session.error_file.name}"'
        return response
        
    except UploadSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
